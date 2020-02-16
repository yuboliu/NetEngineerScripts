# -*- coding: UTF-8 -*-
# @Time    : 2020/1/2 13:09
# @Author  : 刘钰博
# @E-mail  : lolipop.boy@outlook.com

import sys
import os
import openpyxl
import json
import urllib3
import requests
import time
import pysnooper

sys.path.append(os.path.abspath("../"))
import lib.ISEsettings

ersUsername = lib.ISEsettings.ersUsername
ersPassword = lib.ISEsettings.ersPassword
host = lib.ISEsettings.host
xlsx_path = lib.ISEsettings.xlsx_path
sheet_name = lib.ISEsettings.internal_sheet_name
auth = lib.ISEsettings.auth
Headers = lib.ISEsettings.Headers
InternalUserUrl = lib.ISEsettings.InternalUserUrl
IdentityGroupURL = lib.ISEsettings.IdentityGroupURL
IdentityGroupSearchURL = lib.ISEsettings.IdentityGroupSearchURL
the_operation = lib.ISEsettings.the_operation


def before_end(func):
    def decorator(*args, **kw):
        start_time = time.time()
        urllib3.disable_warnings()
        func(*args, **kw)
        local_time = time.time() - start_time
        print("程序执行耗时: {:.2f}".format(local_time))
    return decorator


# @pysnooper.snoop(prefix="generated_data: ")
def generated_data(sheet):
    wb = openpyxl.load_workbook(xlsx_path)
    wb_sheet = wb[sheet]
    max_column = wb_sheet.max_column
    max_row = wb_sheet.max_row
    row_data_list = []
    row_data = {}
    for i in range(2, max_row + 1):
        row_data.setdefault(wb_sheet.cell(1, i).value)
        for j in range(1, max_column + 1):
            row_data[wb_sheet.cell(1, j).value] = wb_sheet.cell(i, j).value
        row_data_list.append(row_data.copy())
    wb.close()
    return row_data_list


def echo_result(name, r):
    def code200():
        print("修改成功: {0}".format(name))

    def code201():
        print("添加成功: {0}".format(name))

    def code204():
        print("删除成功: {0}".format(name))

    def other():
        print(" 操作失败: {2} 错误代码: {0}, 错误返回: {1}".format(r.status_code, r.text, name))
        raise Exception

    def switch(x):
        return {
            200: code200,
            201: code201,
            204: code204
        }.get(x, other)
    switch(r.status_code)()


# @pysnooper.snoop(prefix="add_code: ")
@before_end
def create(sheet):
    row_data_list = generated_data(sheet)
    for data_dict in row_data_list:
        name = "{0}{1}".format(str(data_dict['用户名']), str(data_dict['用户名后缀']))
        password = data_dict['密码']
        identity_groups = str(data_dict['用户组']).split(',')
        r_sess = requests.session()
        identity_group_list = []
        for identity_group in identity_groups:
            r = r_sess.get("{}/{}".format(IdentityGroupSearchURL, identity_group), auth=auth, headers=Headers, verify=False, timeout=(3, 5))
            r_dict = json.loads(r.text)
            identity_group_list.append(r_dict['IdentityGroup']['id'])
        if len(identity_group_list) > 1:
            identity_group_str = ",".join(identity_group_list)
        else:
            identity_group_str = "".join(identity_group_list)
        custom_attributes = str(data_dict['自定义属性']).split(',')
        payload = {
            "InternalUser": {
                "name": name,
                "password": password,
                "enabled": "true",
                "changePassword": "false",
                "passwordIDStore": "Internal Users",
                "identityGroups": identity_group_str,
                "customAttributes": {
                    custom_attributes[0]: data_dict["IMSI号"],
                    custom_attributes[1]: data_dict["广域口 IP"]
                }
            }
        }
        payload_json = json.dumps(payload)
        r = r_sess.post(InternalUserUrl, auth=auth, data=payload_json, headers=Headers, verify=False)
        echo_result(name, r)


@before_end
def action(sheet, operation):
    row_data_list = generated_data(sheet)
    for data_dict in row_data_list:
        name = "{0}{1}".format(str(data_dict['用户名']), str(data_dict['用户名后缀']))
        r_sess = requests.session()
        r = r_sess.get("{0}?filter=name.EQ.{1}".format(InternalUserUrl, name), auth=auth, headers=Headers, verify=False)
        r_dict = json.loads(r.text)
        resource_list = (r_dict["SearchResult"]["resources"])
        context_dict = resource_list[0]
        internal_user_id = context_dict["id"]

        def change_status():
            payload = {
                "InternalUser": {
                    "id": internal_user_id,
                    "name": name,
                    "enabled": operation
                }
            }
            payload_json = json.dumps(payload)
            r = r_sess.put("{0}/{1}".format(InternalUserUrl, internal_user_id), auth=auth, data=payload_json, headers=Headers, verify=False)
            echo_result(name, r)

        def delete():
            r = r_sess.delete("{0}/{1}".format(InternalUserUrl, internal_user_id), auth=auth, headers=Headers, verify=False)
            echo_result(name, r)

        def other():
            print("the_operation 输入错误")
            raise Exception

        def switch(x):
            return {
                "true": change_status,
                "false": change_status,
                "delete": delete
            }.get(x, other)
        switch(the_operation)()


if __name__ == "__main__":
    # create(sheet_name)
    action(sheet_name, the_operation)
