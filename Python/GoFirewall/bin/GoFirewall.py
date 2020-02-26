#!/usr/bin/env python
# -*- coding: UTF-8 -*-
"""
@Author: 刘钰博
@Contact: lolipop.boy@outlook.com
@Time: 2020/2/26 9:24
"""

import sys
import os
import openpyxl

sys.path.append(os.path.abspath("../"))
import lib.ConfigSettings
XLSX_Path = lib.ConfigSettings.XLSX_Path
DataDict = {}


def log(level):
    import logging
    import datetime
    today = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")
    go_level = {"info": logging.INFO, "error": logging.ERROR, "debug": logging.DEBUG, "warning": logging.WARNING,
                "critical": logging.CRITICAL}
    logger = logging.getLogger()
    logger.setLevel(go_level[level])
    fh = logging.FileHandler('..\\log\\{0}.log'.format(today))
    ch = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s %(filename)s[line:%(lineno)d] %(levelname)s %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


def oper_excel(sheet_name):
    wb = openpyxl.load_workbook(XLSX_Path)
    wb_sheet = wb[sheet_name]
    max_column = wb_sheet.max_column
    max_row = wb_sheet.max_row
    for i in range(1, max_column + 1):
        DataDict.setdefault(wb_sheet.cell(1, i).value)
    for i in range(2, max_row + 1):
        for j in range(1, max_column + 1):
            DataDict[wb_sheet.cell(1, j).value] = wb_sheet.cell(i, j).value
        logger.debug(DataDict)
        generate_command()


def generate_command():
    if DataDict['Behavior'] == "Double":
        logger.debug("{} = Double".format(DataDict['Behavior']))
    elif DataDict['Behavior'] == "Inbound":
        logger.debug("{} = Inbound".format(DataDict['Behavior']))
    elif DataDict['Behavior'] == "Outbound":
        logger.debug("{} = Outbound".format(DataDict['Behavior']))
    else:
        logger.error("{} is unknown value".format(DataDict['Behavior']))
        raise ValueError(DataDict['Behavior'])


if __name__ == "__main__":
    logger = log("debug")
    oper_excel("Sheet1")

