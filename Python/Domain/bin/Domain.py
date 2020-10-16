# -*- coding: UTF-8 -*-
# @Time    : 2020/1/3 13:09
# @Author  : LiuYuBo
# @Software: PyCharmCM

import sys
import os
import openpyxl
import datetime
from multiprocessing import Pool
import time
from ftplib import FTP

sys.path.append(os.path.abspath("../"))
sys.path.append(os.path.abspath("../../"))
import lib.BackupConfigSettings
from CommonLib import log as log

XLSX_Path = lib.BackupConfigSettings.XLSX_Path
DataDict = lib.BackupConfigSettings.DataDict
Global_ErrorMess = lib.BackupConfigSettings.Global_ErrorMess
FtpServerIP = lib.BackupConfigSettings.FtpServerIP
FtpServerUsername = lib.BackupConfigSettings.FtpServerUsername
FtpServerPassword = lib.BackupConfigSettings.FtpServerPassword
FtpServerRoot = lib.BackupConfigSettings.FtpServerRoot
H3C = lib.BackupConfigSettings.H3C
MaiPu = lib.BackupConfigSettings.MaiPu
Cisco = lib.BackupConfigSettings.Cisco
run_cmd_dict = lib.BackupConfigSettings.run_cmd_dict
H3C_Cv7 = lib.BackupConfigSettings.H3C_Cv7

today = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")


def gen_data(max_row, max_column, wb_sheet):
    data_list = []
    for i in range(1, max_column + 1):
        DataDict.setdefault(wb_sheet.cell(1, i).value)
    for i in range(2, max_row + 1):
        for j in range(1, max_column + 1):
            DataDict[wb_sheet.cell(1, j).value] = wb_sheet.cell(i, j).value
        data_list.append(DataDict.copy())
    logger.info(data_list)
    iter_datalist = iter(data_list)
    return iter_datalist


# def process_go(iter_datalist):
def process_go(iter_data):
    import subprocess
    import re
    res_data = []
    while True:
        try:
            data = next(iter_data)
        except StopIteration:
            break
        else:
            # cmd = 'curl http://www.cip.cc/58.83.128.0'
            cmd = "curl http://www.cip.cc/{0}".format(data['IP段'])
            try:
                res = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=10)
            except subprocess.TimeoutExpired as e:
                time.sleep(60)
            else:
                if res.returncode == 0:
                    # print(bytes.decode(res.stdout))
                    str_res = bytes.decode(res.stdout)
                    re_regex = re.compile(r'数据二.*?(?=URL)', re.DOTALL)
                    re_res = re_regex.findall(str_res)
                    re_res = str(re_res).replace("\\t", "").replace("['","").replace("']","").replace(" | ", "_").replace("\\n", "  ")
                    data['百度查询结果'] = str(re_res)
                    print(data)
                    res_data.append(data.copy())
                # print(res_data)
            finally:
                time.sleep(8)
    return res_data


def main(sheet_name):
    wb = openpyxl.load_workbook(XLSX_Path)
    wb_sheet = wb[sheet_name]
    max_column = wb_sheet.max_column
    max_row = wb_sheet.max_row
    iter_data = gen_data(max_row, max_column, wb_sheet)
    res = process_go(iter_data)
    wb.close()
    data_to_excel(res)


def data_to_excel(res):
    import pandas as pd
    df = pd.DataFrame.from_dict(res)
    write = pd.ExcelWriter(r'../result/Result.xlsx')
    df.to_excel(r'../result/Result.xlsx')


if __name__ == "__main__":
    logger = log("info")
    main("Sheet1")
