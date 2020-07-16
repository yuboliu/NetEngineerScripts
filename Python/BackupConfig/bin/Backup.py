# -*- coding: UTF-8 -*-
# @Time    : 2020/1/3 13:09
# @Author  : LiuYuBo
# @Software: PyCharmCM

import sys
import os
import openpyxl
import datetime
import paramiko
import json
from multiprocessing import Pool
import time
from ftplib import FTP

sys.path.append(os.path.abspath("../"))
sys.path.append(os.path.abspath("../../"))
import lib.BackupConfigSettings
from CommonLib import log as log

XLSX_Path = lib.BackupConfigSettings.XLSX_Path
DataDict = lib.BackupConfigSettings.DataDict
Global_ErrorMess = lib.BackupConfigSettings.Global_ErrorMess
FtpServerIP = lib.BackupConfigSettings.FtpServerIP
FtpServerUsername = lib.BackupConfigSettings.FtpServerUsername
FtpServerPassword = lib.BackupConfigSettings.FtpServerPassword
FtpServerRoot = lib.BackupConfigSettings.FtpServerRoot
H3C = lib.BackupConfigSettings.H3C
MaiPu = lib.BackupConfigSettings.MaiPu
Cisco = lib.BackupConfigSettings.Cisco
run_cmd_dict = lib.BackupConfigSettings.run_cmd_dict
H3C_Cv7 = lib.BackupConfigSettings.H3C_Cv7

today = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M")


def test_fun():
    import random
    r = random.randint(1, 5)
    time.sleep(r)
    return r, __name__


def common_ssh(DataDict):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(hostname=DataDict['地址'], port=22, username=DataDict['用户名'], password=DataDict['密码'],
                    look_for_keys=False, timeout=3)
    except Exception as r:
        raise "设备名称: {0} 设备地址: {1} 错误: {2}".format(DataDict['设备名称'], DataDict['地址'], r)
    else:
        if DataDict['设备厂商'] == "H3C-v7":
            stdin, stdout, stderr = ssh.exec_command(
                H3C_Cv7.format(FtpServerUsername, FtpServerPassword, FtpServerIP, DataDict['设备名称'], today))
            b_result = stdout.read()
            r = bytes.decode(b_result)
            print("设备名称: {0} 设备地址: {1}".format(DataDict['设备名称'], DataDict['地址']))
            return r

        else:
            channel = ssh.invoke_shell(height=999)
            stdin = channel.makefile('wb')
            stdout = channel.makefile('rb')
            run_cmd = DataDict['设备厂商']
            stdin.write(
                run_cmd_dict[run_cmd].format(FtpServerIP, FtpServerUsername, FtpServerPassword, DataDict['设备名称'], today,
                                             DataDict['地址'], DataDict['Enable密码']))
            r = stdout.read().decode('gbk')
            print("设备名称: {0} 设备地址: {1}".format(DataDict['设备名称'], DataDict['地址']))
            return r
    finally:
        ssh.close()


def make_dir(wb_sheet, max_column, max_row):
    all_device_vendor = []
    for i in range(1, max_column + 1):
        DataDict.setdefault(wb_sheet.cell(1, i).value)
        if wb_sheet.cell(1, i).value == "设备厂商":
            DeviceVendor_ColNum = i
    for i in range(2, max_row + 1):
        all_device_vendor.append(wb_sheet.cell(i, DeviceVendor_ColNum).value)
    all_device_vendor = set(all_device_vendor)
    for n, i in enumerate(all_device_vendor):
        if i == "H3C-v7":
            all_device_vendor[n] = "H3C"
    # all_device_vendor.discard("H3C-v7")
    logger.info(all_device_vendor)
    for i in all_device_vendor:
        os.makedirs(("{}\\{}\\{}".format(FtpServerRoot, today, i)))


def ftp_up(filename):
    ftp = FTP()
    ftp.set_debuglevel(2)
    ftp.connect(FtpServerIP)
    ftp.login(FtpServerUsername, FtpServerPassword)
    print(ftp.getwelcome())
    # ftp.cwd('{0}\\{1}'.format(FtpServerRoot, today))
    ftp.cwd(today)
    ftp.mkd(filename)
    ftp.set_debuglevel(0)
    ftp.quit()
    logger.info("ftp up test ok")


def gen_data(max_row, max_column, wb_sheet):
    data_list=[]
    for i in range(1, max_column + 1):
        DataDict.setdefault(wb_sheet.cell(1, i).value)
    for i in range(2, max_row + 1):
        for j in range(1, max_column + 1):
            DataDict[wb_sheet.cell(1, j).value] = wb_sheet.cell(i, j).value
        data_list.append(DataDict.copy())
    logger.info(data_list)
    iter_datalist = iter(data_list)
    return iter_datalist


def process_go(iter_datalist):
    exit_flag = False
    results = []
    p_pool = Pool(8)
    while True:
        for k in range(8):
            try:
                data = next(iter_datalist)
            except StopIteration:
                exit_flag = True
                break
            logger.info(data)
            res = p_pool.apply_async(common_ssh, args=(data,))
            # res = p_pool.apply_async(test_fun)
            results.append(res)
        if exit_flag:
            break
    p_pool.close()
    p_pool.join()
    for result in results:
        # print(result.get())
        logger.info(result.get())


def main(sheet_name):
    global Global_ErrorMess
    wb = openpyxl.load_workbook(XLSX_Path)
    wb_sheet = wb[sheet_name]
    max_column = wb_sheet.max_column
    max_row = wb_sheet.max_row
    make_dir(wb_sheet, max_column, max_row)
    ftp_up('TestDir')
    gen_date = gen_data(max_row, max_column, wb_sheet)
    process_go(gen_date)
    wb.close()


if __name__ == "__main__":
    logger = log("info")
    main("WorkNet")
